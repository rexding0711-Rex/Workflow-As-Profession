#!/usr/bin/env python3
"""
xlsx_tools.py — Formula recalculation and verification for Excel files.

Usage:
  python3 xlsx_tools.py recalc <file.xlsx>     Recalculate all formulas
  python3 xlsx_tools.py verify <file.xlsx>      Verify no formula errors remain
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import sys

# Forbidden functions (incompatible with Excel 2019 and earlier)
FORBIDDEN_FUNCTIONS = [
    'FILTER', 'UNIQUE', 'SORT', 'SORTBY', 'XLOOKUP', 'XMATCH',
    'SEQUENCE', 'LET', 'LAMBDA', 'RANDARRAY'
]

# Implicit array formula pattern
IMPLICIT_ARRAY_PATTERN = re.compile(r'MATCH\s*\(\s*TRUE\s*\(\s*\)', re.IGNORECASE)


def cmd_recalc(args):
    """Recalculate all formulas in an Excel file."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(json.dumps({"status": "error", "message": f"File not found: {filepath}"}))
        sys.exit(2)

    # Option 1: formulas library
    try:
        import formulas
        print("Recalculating with formulas library...")
        xl_model = formulas.ExcelModel().loads(filepath).finish()
        xl_model.calculate()

        # Write to temp file for verification only (formulas strips charts/styles)
        verify_path = filepath.replace('.xlsx', '_recalc_verify.xlsx')
        xl_model.write(dirpath=verify_path)
        print(json.dumps({
            "status": "success",
            "method": "formulas",
            "note": f"Verification file: {verify_path} (do NOT deliver this — it strips charts/styles). Deliver the original file.",
        }))

        # Set calcMode=auto on the ORIGINAL file so Excel recalculates on open
        _set_calc_mode_auto(filepath)
        return
    except ImportError:
        print("formulas library not available, trying libreoffice...")
    except Exception as e:
        print(f"formulas library failed: {e}, trying libreoffice...")

    # Option 2: libreoffice headless
    if shutil.which('libreoffice'):
        try:
            outdir = os.path.dirname(os.path.abspath(filepath)) or '.'
            subprocess.run([
                'libreoffice', '--headless', '--calc', '--convert-to', 'xlsx',
                '--outdir', outdir, filepath
            ], check=True, capture_output=True, timeout=120)
            print(json.dumps({"status": "success", "method": "libreoffice"}))
            return
        except Exception as e:
            print(f"libreoffice failed: {e}, falling back to calcMode=auto...")

    # Fallback: set calcMode=auto
    _set_calc_mode_auto(filepath)
    print(json.dumps({
        "status": "success",
        "method": "calcMode_auto",
        "note": "Neither formulas nor libreoffice available. Set calcMode=auto — Excel/WPS will recalculate on open.",
    }))


def _set_calc_mode_auto(filepath):
    """Set calcMode to auto so Excel recalculates on open."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        if wb.calculation is None:
            from openpyxl.workbook.properties import CalcProperties
            wb.calculation = CalcProperties()
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.save(filepath)
    except Exception as e:
        print(f"Warning: could not set calcMode: {e}", file=sys.stderr)


def cmd_verify(args):
    """Verify an Excel file for formula errors, forbidden functions, and implicit arrays."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(json.dumps({"status": "error", "message": f"File not found: {filepath}"}))
        sys.exit(2)

    from openpyxl import load_workbook

    issues = []

    # Check 1: Formula errors (open with data_only to see cached values)
    try:
        wb_data = load_workbook(filepath, data_only=True)
        for ws in wb_data.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('#'):
                        issues.append({
                            "type": "formula_error",
                            "severity": "error",
                            "location": f"{ws.title}!{cell.coordinate}",
                            "value": cell.value,
                        })
    except Exception as e:
        issues.append({"type": "load_error", "severity": "error", "detail": str(e)})

    # Check 2: Forbidden functions + implicit arrays (open with formulas)
    try:
        wb_formulas = load_workbook(filepath)
        for ws in wb_formulas.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value.upper()

                        # Forbidden functions
                        for func in FORBIDDEN_FUNCTIONS:
                            if func + '(' in formula:
                                issues.append({
                                    "type": "forbidden_function",
                                    "severity": "warning",
                                    "location": f"{ws.title}!{cell.coordinate}",
                                    "function": func,
                                    "formula": cell.value[:80],
                                })

                        # Implicit array formula
                        if IMPLICIT_ARRAY_PATTERN.search(cell.value):
                            issues.append({
                                "type": "implicit_array",
                                "severity": "warning",
                                "location": f"{ws.title}!{cell.coordinate}",
                                "formula": cell.value[:80],
                                "hint": "Use SUMPRODUCT or helper column instead of MATCH(TRUE(), ...)",
                            })
    except Exception as e:
        issues.append({"type": "load_error", "severity": "error", "detail": str(e)})

    # Output
    errors = [i for i in issues if i['severity'] == 'error']
    warnings = [i for i in issues if i['severity'] == 'warning']

    result = {
        "status": "pass" if not errors else "fail",
        "file": filepath,
        "error_count": len(errors),
        "warning_count": len(warnings),
        "issues": issues,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0 if not errors else 1)


def main():
    parser = argparse.ArgumentParser(description="Excel formula tools")
    subparsers = parser.add_subparsers(dest="command")

    recalc_p = subparsers.add_parser("recalc", help="Recalculate all formulas")
    recalc_p.add_argument("file", help="Excel file path")
    recalc_p.set_defaults(func=cmd_recalc)

    verify_p = subparsers.add_parser("verify", help="Verify formula errors and forbidden functions")
    verify_p.add_argument("file", help="Excel file path")
    verify_p.set_defaults(func=cmd_verify)

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        sys.exit(0)
    args.func(args)


if __name__ == "__main__":
    main()
